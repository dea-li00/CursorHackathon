import os
from datetime import datetime
from decimal import Decimal
from typing import List, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from ..models import APBill, WorkflowStatus
from ..config import settings


class ExcelExporter:
    def __init__(self):
        self.export_dir = settings.EXPORT_DIR
        os.makedirs(self.export_dir, exist_ok=True)
    
    def export_bills(self, bills: List[APBill], filename: Optional[str] = None) -> str:
        """Export bills to Excel file"""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"AP_Bills_{timestamp}.xlsx"
        
        filepath = os.path.join(self.export_dir, filename)
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create headers sheet
        headers_df = self._create_headers_dataframe(bills)
        self._create_headers_sheet(wb, headers_df)
        
        # Create lines sheet
        lines_df = self._create_lines_dataframe(bills)
        self._create_lines_sheet(wb, lines_df)
        
        # Save workbook
        wb.save(filepath)
        return filepath
    
    def _create_headers_dataframe(self, bills: List[APBill]) -> pd.DataFrame:
        """Create headers dataframe"""
        headers_data = []
        
        for bill in bills:
            headers_data.append({
                "id": bill.id,
                "vendor_name": bill.vendor_name,
                "vendor_tax_id": bill.vendor_tax_id or "",
                "vendor_iban": bill.vendor_iban or "",
                "invoice_number": bill.invoice_number,
                "invoice_date": bill.invoice_date.strftime("%Y-%m-%d") if bill.invoice_date else "",
                "due_date": bill.due_date.strftime("%Y-%m-%d") if bill.due_date else "",
                "po_number": bill.po_number or "",
                "currency": bill.currency,
                "subtotal": float(bill.subtotal) if bill.subtotal else 0.0,
                "tax": float(bill.tax) if bill.tax else 0.0,
                "tax_rate_pct": float(bill.tax_rate_pct) if bill.tax_rate_pct else 0.0,
                "total": float(bill.total) if bill.total else 0.0,
                "status": bill.workflow.status.value,
                "notes": bill.notes or "",
                "created_at": bill.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                "updated_at": bill.updated_at.strftime("%Y-%m-%d %H:%M:%S")
            })
        
        return pd.DataFrame(headers_data)
    
    def _create_lines_dataframe(self, bills: List[APBill]) -> pd.DataFrame:
        """Create lines dataframe"""
        lines_data = []
        
        for bill in bills:
            for i, line in enumerate(bill.lines):
                lines_data.append({
                    "bill_id": bill.id,
                    "line_index": i + 1,
                    "description": line.description,
                    "quantity": float(line.quantity) if line.quantity else None,
                    "unit_price": float(line.unit_price) if line.unit_price else None,
                    "amount": float(line.amount) if line.amount else 0.0,
                    "sku": line.sku or "",
                    "cost_center": line.cost_center or "",
                    "gl_account": line.gl_account or "",
                    "tax_code": line.tax_code or ""
                })
        
        return pd.DataFrame(lines_data)
    
    def _create_headers_sheet(self, wb: Workbook, df: pd.DataFrame) -> None:
        """Create and format headers sheet"""
        ws = wb.create_sheet("AP_Bill_Headers")
        
        # Add data to worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Format headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Format currency columns
        currency_columns = ["subtotal", "tax", "total"]
        for col_idx, col_name in enumerate(df.columns, 1):
            if col_name in currency_columns:
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    if cell.value is not None:
                        cell.number_format = "$#,##0.00"
        
        # Format percentage column
        if "tax_rate_pct" in df.columns:
            tax_rate_col = df.columns.get_loc("tax_rate_pct") + 1
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=tax_rate_col)
                if cell.value is not None:
                    cell.number_format = "0.00%"
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
    
    def _create_lines_sheet(self, wb: Workbook, df: pd.DataFrame) -> None:
        """Create and format lines sheet"""
        ws = wb.create_sheet("AP_Bill_Lines")
        
        # Add data to worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Format headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Format currency columns
        currency_columns = ["quantity", "unit_price", "amount"]
        for col_idx, col_name in enumerate(df.columns, 1):
            if col_name in currency_columns:
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    if cell.value is not None:
                        cell.number_format = "$#,##0.00"
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
    
    def get_export_summary(self, bills: List[APBill]) -> dict:
        """Get export summary statistics"""
        total_bills = len(bills)
        total_amount = sum(float(bill.total) for bill in bills if bill.total)
        
        status_counts = {}
        for bill in bills:
            status = bill.workflow.status.value
            status_counts[status] = status_counts.get(status, 0) + 1
        
        currency_counts = {}
        for bill in bills:
            currency = bill.currency
            currency_counts[currency] = currency_counts.get(currency, 0) + 1
        
        return {
            "total_bills": total_bills,
            "total_amount": total_amount,
            "status_counts": status_counts,
            "currency_counts": currency_counts,
            "export_timestamp": datetime.now().isoformat()
        }
